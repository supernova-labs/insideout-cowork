"""
grid_library — grid editorial mensal por marca (InsideOut).

Terceiro pilar da geração de social media (os outros dois: `style_library`
= "como a peça parece"; `product_library` = "o que é o produto e como a marca
fala"). Aqui mora **"o que postar e quando"**: um grid = 1 marca × 1 mês, com
o esqueleto semanal que a Estela (social media Clinique/EL/TF) usa hoje em
duas planilhas Excel — colapsadas num único artefato canônico.

Princípios (mesmos do `product_library`, ver plano merry-snacking-flame):
- Dados na PASTA DE TRABALHO do usuário — 1 JSON por grid
  (`<workspace>/grids/<marca>/<AAAA-MM>.json`). Nunca no plugin dir.
- O grid HTML é SEMPRE derivado (gerado do JSON) — nunca fonte de dado, nunca
  editado por string (a planilha apodrecia exatamente por edição à mão).
- Escrita atômica; delete é soft (.trash/); nada de "apagar tudo".
- Sem workspace, cai no seed embarcado (read-only) — zero-config funciona.
- **Regras da Estela** vivem em `grids/rules/<marca>.md` (Markdown editável
  por humano, não hardcoded) e o **calendário comemorativo** em
  `grids/calendar/<ano>.md` (compartilhado entre marcas). Consumidos na Fase 2
  (`generate_from_briefing`); a Fase 1 só os materializa do seed.

A disciplina UWP-safe (nunca `.resolve()` em `__file__` de plugin; desconfiar
de todo stat do plugin dir) vive em `_libcommon`. Sem efeitos colaterais no
import.
"""
from __future__ import annotations

import calendar
import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path

import _libcommon as lc

SCHEMA_VERSION = 1
LIB_DIRNAME = "grids"
ENV_OVERRIDE = "GRIDS_DIR"

_CORE_DIR = lc.CORE_DIR
_SEED_FILE = _CORE_DIR / "grids.seed.json"
_RULES_SEED_DIR = _CORE_DIR / "rules-seed"
_CALENDAR_SEED_DIR = _CORE_DIR / "calendar-seed"
_TEMPLATE_FILE = _CORE_DIR / "grid-template.html"
_PLACEHOLDER = "/*__GRID_JSON__*/null"

# Subdirs reservados dentro de grids/ — NÃO são marcas.
_RESERVED = {"rules", "calendar", "mockups", ".trash"}

# Abreviação PT do dia da semana, indexada com domingo=0 (a planilha da Estela
# é DOMINGO..SÁBADO).
_DOW = ["DOM", "SEG", "TER", "QUA", "QUI", "SEX", "SAB"]

_PT_MONTHS = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "março": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8, "setembro": 9,
    "outubro": 10, "novembro": 11, "dezembro": 12,
}

# Escopo de ingestão travado com o Lucas (2026-05-18): só 2026 já valida o
# modelo de dados; histórico 2025 fica fora.
INGEST_YEAR = 2026

# Campos de um "dia" do grid que as operações de post podem mover/editar
# (data e dow são identidade do slot, não conteúdo).
_POST_FIELDS = ("channel", "approach", "product", "subject", "ref",
                "lettering", "mockup", "rationale", "notes")

# Cadência mínima codificada (defaults; a skill pode sobrescrever via kwargs,
# mas o core NÃO parseia número do rules.md — frágil; D4 do plano Fase 2).
MIN_POSTS = 28
MAX_GAP = 2
FOCUS_INTENSITY_DAYS = 3

# Taxonomia de _slot (anota o JULGAMENTO que o agente vai aplicar). O core só
# propõe; a skill decide product/subject/ref/etc. via set_post.
_SLOT_KINDS = ("launch_anchor", "launch_intensity", "calendar_hook",
               "focus_intercalation", "hero_fill", "free")


class GridError(lc.LibCommonError):
    """Erro base do grid."""


class GridNotFound(GridError):
    pass


class InvalidGrid(GridError):
    pass


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def slugify(name: str) -> str:
    return lc.slugify(name, fallback="marca")


def _subdirs(lib_dir: Path):
    """Dirs/arquivos fixos do grid store. As pastas por-marca são dinâmicas
    (uma por slug de marca) e ficam fora desta tupla."""
    return (lib_dir / "rules", lib_dir / "calendar", lib_dir / "mockups",
            lib_dir / ".trash", lib_dir / f"{LIB_DIRNAME}.html")


def _ensure_dirs(lib_dir: Path) -> Path:
    rules, cal, mock, trash, _ = _subdirs(lib_dir)
    lc.ensure_dirs(lib_dir, rules, cal, mock, trash)
    return lib_dir


def find_library_dir(start: Path | None = None,
                     create: bool = True) -> Path | None:
    """
    Resolve o diretório dos grids. Ordem:
      1. $GRIDS_DIR (se setada);
      2. busca pra cima a partir de `start` (default cwd) por um `grids/`
         existente (para na raiz do git/filesystem);
      3. se nada e create=True: cria `<cwd>/grids/`. create=False: None.
    """
    d = lc.find_library_dir(LIB_DIRNAME, ENV_OVERRIDE, start, create)
    if d is not None and create:
        return _ensure_dirs(d)
    return d


def _norm_month(month, year: int | None = None) -> str:
    """Normaliza mês para 'AAAA-MM'. Aceita: 'AAAA-MM', 'MM', 'M', ou nome PT
    ('maio'); year usado quando o mês vem sem ano."""
    s = str(month).strip().lower()
    m = re.fullmatch(r"(\d{4})-(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    if s in _PT_MONTHS:
        mm = _PT_MONTHS[s]
    elif s.isdigit() and 1 <= int(s) <= 12:
        mm = int(s)
    else:
        raise InvalidGrid(
            f"Mês '{month}' inválido — use 'AAAA-MM', número 1-12 ou nome "
            f"(ex.: 'maio').")
    if year is None:
        raise InvalidGrid(
            f"Mês '{month}' sem ano — informe year ou use 'AAAA-MM'.")
    return f"{int(year):04d}-{mm:02d}"


def _seed() -> dict:
    """Lê grids.seed.json DE FATO (não .exists()). Estrutura: {grids:[...]}."""
    data = lc.read_json_strict(_SEED_FILE)
    if not isinstance(data, dict):
        raise GridError(
            f"grids.seed.json malformado em {_SEED_FILE} — esperado objeto "
            f"com 'grids'.")
    data.setdefault("grids", [])
    return data


# --------------------------------------------------------------------------- #
# bootstrap / lazy-ensure
# --------------------------------------------------------------------------- #
def _grid_path(lib_dir: Path, brand: str, month: str) -> Path:
    return Path(lib_dir) / slugify(brand) / f"{month}.json"


def bootstrap(lib_dir: Path, _render: bool = True) -> dict:
    """
    Semeia o workspace na 1ª vez:
      - grids do seed (1 JSON por marca-mês) se nenhum grid existe ainda;
      - regras por marca (`rules-seed/*.md` -> `rules/`), só as ausentes;
      - calendário comemorativo (`calendar-seed/*.md` -> `calendar/`), idem.
    Idempotente: nunca sobrescreve arquivo existente. Regras e calendário são
    copiados SEMPRE que faltam (responsabilidade distinta dos grids — mesma
    lógica de thumbnails/fotos no style/product_library).
    Retorna {grids, rules:{...}, calendar:{...}}.
    """
    lib_dir = _ensure_dirs(Path(lib_dir))
    rules_dir, cal_dir, _, _, _ = _subdirs(lib_dir)

    # Validar seed lendo DE FATO (não .exists()).
    try:
        _seed()
    except (OSError, ValueError) as e:
        raise GridError(
            f"grids.seed.json ilegível em {_CORE_DIR} ({e!r}) — plugin mal "
            f"empacotado ou ambiente bloqueando acesso a arquivo.") from e

    # Regras + calendário: disciplina anti-stat-mentiroso do _libcommon.
    # raise_on_empty=False: o seed PODE legitimamente ter 0 arquivos numa
    # instalação enxuta; o que não pode é o os.listdir do plugin dir falhar
    # silenciosamente (isso copy_seed_assets levanta de qualquer forma).
    rules_res = lc.copy_seed_assets(
        _RULES_SEED_DIR, rules_dir, exts={".md"},
        error_cls=GridError, label="rules-seed", raise_on_empty=False)
    cal_res = lc.copy_seed_assets(
        _CALENDAR_SEED_DIR, cal_dir, exts={".md"},
        error_cls=GridError, label="calendar-seed", raise_on_empty=False)

    # Guard de idempotência: grids só são semeados uma vez.
    if _has_any_grid(lib_dir):
        if _render:
            render_grids(lib_dir)
        return {"grids": 0, "rules": rules_res, "calendar": cal_res}

    seed = _seed()
    ng = 0
    for g in seed.get("grids", []):
        b, m = g.get("brand"), g.get("month")
        if not b or not m:
            continue
        dest = _grid_path(lib_dir, b, m)
        if dest.exists():
            continue
        dest.parent.mkdir(parents=True, exist_ok=True)
        lc.atomic_write(dest, json.dumps(g, ensure_ascii=False, indent=2) + "\n")
        ng += 1
    if _render:
        render_grids(lib_dir)
    return {"grids": ng, "rules": rules_res, "calendar": cal_res}


def _ensure_ready(lib_dir: Path | None = None) -> Path:
    """Lazy-ensure: resolve a pasta (criando se preciso) e roda bootstrap
    (idempotente, sem render). Workspace vira fonte única ANTES de
    exibir/curar/mutar — elimina o fallback-fantasma de seed."""
    lib_dir = _ensure_dirs(
        find_library_dir() if lib_dir is None else Path(lib_dir))
    bootstrap(lib_dir, _render=False)
    return lib_dir


# --------------------------------------------------------------------------- #
# leitura (pura: zero-config, cai no seed)
# --------------------------------------------------------------------------- #
def _brand_dirs(lib_dir: Path):
    for child in sorted(Path(lib_dir).iterdir()):
        if child.is_dir() and child.name not in _RESERVED:
            yield child


def _has_any_grid(lib_dir: Path) -> bool:
    for bd in _brand_dirs(lib_dir):
        if any(bd.glob("*.json")):
            return True
    return False


def _read_grids(lib_dir: Path) -> list[dict]:
    out: list[dict] = []
    for bd in _brand_dirs(lib_dir):
        for fp in sorted(bd.glob("*.json")):
            try:
                out.append(json.loads(fp.read_text(encoding="utf-8")))
            except (json.JSONDecodeError, OSError):
                continue
    return out


def _resolve(lib_dir) -> tuple[list[dict], str]:
    if lib_dir is None:
        lib_dir = find_library_dir(create=False)
    if lib_dir is not None:
        items = _read_grids(Path(lib_dir))
        if items:
            return items, "workspace"
    return _seed().get("grids", []), "seed"


def _sort_key(g: dict):
    return (g.get("brand", ""), g.get("month", ""))


def list_grids(brand=None, lib_dir: Path | None = None) -> list[dict]:
    """Lista grids (resumo). Filtra por marca (slug/nome) se dado."""
    items, _ = _resolve(lib_dir)
    if brand is not None:
        bslug = slugify(brand)
        items = [g for g in items if slugify(g.get("brand", "")) == bslug]
    return [
        {"brand": g.get("brand"), "month": g.get("month"),
         "focusProducts": g.get("focusProducts", []),
         "posts": sum(len(w.get("days", [])) for w in g.get("weeks", []))}
        for g in sorted(items, key=_sort_key)
    ]


def get_grid(brand, month, lib_dir: Path | None = None) -> dict:
    """Grid completo de uma marca-mês. month aceita 'AAAA-MM' ou nome/num
    (precisa do ano embutido ou via 'AAAA-MM')."""
    items, _ = _resolve(lib_dir)
    bslug = slugify(brand)
    try:
        m = _norm_month(month)
    except InvalidGrid:
        m = str(month).strip()
    for g in items:
        if slugify(g.get("brand", "")) == bslug and g.get("month") == m:
            return g
    raise GridNotFound(
        f"Grid '{brand}/{month}' não encontrado (use list_grids / --list).")


# --------------------------------------------------------------------------- #
# construção de esqueleto (calendário do mês, domingo-primeiro)
# --------------------------------------------------------------------------- #
def _empty_day(date_iso: str) -> dict:
    d = datetime.strptime(date_iso, "%Y-%m-%d")
    # weekday(): seg=0..dom=6 -> nosso índice dom=0..sab=6
    dow_idx = (d.weekday() + 1) % 7
    return {
        "date": date_iso,
        "dow": _DOW[dow_idx],
        "channel": None,
        "approach": None,
        "product": None,
        "subject": None,
        "ref": None,
        "lettering": {},
        "mockup": None,
        "rationale": "",
        "notes": "",
    }


def _build_weeks(year: int, month: int) -> list[dict]:
    """Esqueleto de semanas domingo→sábado, só os dias do próprio mês
    (espelha a planilha: dias de outros meses ficam vazios/ausentes)."""
    cal = calendar.Calendar(firstweekday=6)  # 6 = domingo
    weeks: list[dict] = []
    for i, week in enumerate(cal.monthdatescalendar(year, month), start=1):
        days = [_empty_day(d.isoformat()) for d in week if d.month == month]
        if days:
            weeks.append({"n": i, "days": days})
    # renumera sequencialmente (semanas sem dia do mês foram puladas)
    for n, w in enumerate(weeks, start=1):
        w["n"] = n
    return weeks


def new_grid(brand: str, month, *, year: int | None = None,
             focus_products: list[str] | None = None,
             lib_dir: Path | None = None, save: bool = True) -> dict:
    """Cria um grid vazio (esqueleto semanal do mês) pra preencher
    conversacionalmente ou via ingestão. month: 'AAAA-MM' ou núm/nome+year."""
    m = _norm_month(month, year)
    yy, mm = int(m[:4]), int(m[5:7])
    grid = {
        "schemaVersion": SCHEMA_VERSION,
        "brand": slugify(brand),
        "month": m,
        "focusProducts": list(focus_products or []),
        "weeks": _build_weeks(yy, mm),
        "createdAt": lc.now(),
        "updatedAt": lc.now(),
    }
    if save:
        return save_grid(grid, lib_dir=lib_dir)
    return grid


# --------------------------------------------------------------------------- #
# escrita
# --------------------------------------------------------------------------- #
def _validate_grid(grid: dict) -> None:
    if not isinstance(grid, dict):
        raise InvalidGrid("grid não é um objeto.")
    if not grid.get("brand"):
        raise InvalidGrid("grid sem 'brand'.")
    if not grid.get("month"):
        raise InvalidGrid("grid sem 'month'.")
    if not re.fullmatch(r"\d{4}-\d{2}", str(grid["month"])):
        raise InvalidGrid(
            f"month '{grid['month']}' fora do formato 'AAAA-MM'.")
    if not isinstance(grid.get("weeks", []), list):
        raise InvalidGrid("grid['weeks'] deve ser lista.")


def save_grid(grid: dict, *, lib_dir: Path | None = None) -> dict:
    """Grava o grid canônico (escrita atômica) e regenera o HTML. Fonte única
    — NUNCA editar o HTML; sempre passar por aqui."""
    _validate_grid(grid)
    lib_dir = _ensure_ready(lib_dir)
    grid = dict(grid)
    grid.setdefault("schemaVersion", SCHEMA_VERSION)
    grid["brand"] = slugify(grid["brand"])
    grid.setdefault("createdAt", lc.now())
    grid["updatedAt"] = lc.now()
    dest = _grid_path(lib_dir, grid["brand"], grid["month"])
    dest.parent.mkdir(parents=True, exist_ok=True)
    lc.atomic_write(dest, json.dumps(grid, ensure_ascii=False, indent=2) + "\n")
    render_grids(lib_dir)
    return grid


def delete_grid(brand, month, *, lib_dir: Path | None = None) -> dict:
    """Soft-delete: move o JSON pra .trash/ (recuperável)."""
    lib_dir = _ensure_ready(lib_dir)
    _, _, _, trash_dir, _ = _subdirs(lib_dir)
    m = _norm_month(month) if not re.fullmatch(
        r"\d{4}-\d{2}", str(month)) else str(month)
    fp = _grid_path(lib_dir, brand, m)
    if not fp.is_file():
        raise GridNotFound(f"Grid '{brand}/{month}' não encontrado.")
    trash_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    dest = trash_dir / f"grid.{slugify(brand)}.{m}.{stamp}.json"
    os.replace(fp, dest)
    render_grids(lib_dir)
    return {"deleted": f"{slugify(brand)}/{m}.json", "trash": str(dest)}


# --------------------------------------------------------------------------- #
# operações de post (edição conversacional — reescrevem o JSON, nunca o HTML)
# --------------------------------------------------------------------------- #
def _find_day(grid: dict, date_iso: str):
    for w in grid.get("weeks", []):
        for d in w.get("days", []):
            if d.get("date") == date_iso:
                return d
    raise GridNotFound(
        f"Dia '{date_iso}' não existe no grid {grid.get('brand')}/"
        f"{grid.get('month')}.")


def _content(day: dict) -> dict:
    return {k: day.get(k) for k in _POST_FIELDS}


def _clear(day: dict) -> None:
    for k in _POST_FIELDS:
        day[k] = {} if k == "lettering" else (
            "" if k in ("rationale", "notes") else None)


def move_post(brand, month, src_date: str, dst_date: str, *,
              lib_dir: Path | None = None) -> dict:
    """Move o conteúdo do post de src_date pra dst_date; src fica vazio
    (espelha 'puxa o do dia 10 pra amanhã' da Estela). dst é sobrescrito."""
    grid = get_grid(brand, month, lib_dir)
    src, dst = _find_day(grid, src_date), _find_day(grid, dst_date)
    content = _content(src)
    for k, v in content.items():
        dst[k] = v
    _clear(src)
    return save_grid(grid, lib_dir=lib_dir)


def swap_posts(brand, month, date_a: str, date_b: str, *,
               lib_dir: Path | None = None) -> dict:
    """Troca o conteúdo de dois dias (mantém data/dow de cada slot)."""
    grid = get_grid(brand, month, lib_dir)
    a, b = _find_day(grid, date_a), _find_day(grid, date_b)
    ca, cb = _content(a), _content(b)
    for k in _POST_FIELDS:
        a[k], b[k] = cb[k], ca[k]
    return save_grid(grid, lib_dir=lib_dir)


def set_post(brand, month, date: str, *, lib_dir: Path | None = None,
             **fields) -> dict:
    """Edita campos de um post. Só os de _POST_FIELDS; data/dow são imutáveis."""
    bad = [k for k in fields if k not in _POST_FIELDS]
    if bad:
        raise InvalidGrid(
            f"Campos não editáveis: {', '.join(bad)}. "
            f"Editáveis: {', '.join(_POST_FIELDS)}.")
    grid = get_grid(brand, month, lib_dir)
    day = _find_day(grid, date)
    day.update(fields)
    return save_grid(grid, lib_dir=lib_dir)


def clear_post(brand, month, date: str, *,
                lib_dir: Path | None = None) -> dict:
    """Esvazia um post (slot vira vazio, data/dow preservados)."""
    grid = get_grid(brand, month, lib_dir)
    _clear(_find_day(grid, date))
    return save_grid(grid, lib_dir=lib_dir)


# --------------------------------------------------------------------------- #
# ingestão das planilhas históricas (Fase 1 — só 2026, valida o modelo)
# --------------------------------------------------------------------------- #
def _load_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError as e:
        raise GridError(
            "openpyxl ausente — `pip install -r \"$CORE/requirements.txt\"` "
            "(necessário só para ingest_xlsx).") from e
    return openpyxl


def xlsx_sheets(path: str) -> list[str]:
    """Lista as abas de um .xlsx (pra escolher qual ingerir)."""
    openpyxl = _load_openpyxl()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


_URL_RE = re.compile(r"https?://\S+")
_TELA_RE = re.compile(r"TELA\s*\d+\s*:", re.IGNORECASE)
_LETTER_RE = re.compile(
    r"LETTERING\s+([A-ZÇÃÕÁÉÍÓÚÂÊÔ]+)\s*:\s*(.*)", re.IGNORECASE)


def _parse_story(blob: str) -> dict:
    """Parser pragmático do blob STORY da 'Briefing Design' → lettering/telas/
    refs. Não busca perfeição (Fase 1 valida o modelo); guarda o cru também."""
    text = str(blob or "").strip()
    if not text:
        return {}
    refs = _URL_RE.findall(text)
    telas = [t.strip() for t in _TELA_RE.split(text)[1:]] if _TELA_RE.search(text) else []
    lettering: dict = {}
    for m in _LETTER_RE.finditer(text):
        lettering[m.group(1).strip().lower()] = m.group(2).strip()
    out: dict = {"raw": text}
    if refs:
        out["refs"] = refs
    if telas:
        out["telas"] = telas
    if lettering:
        out.update(lettering)
    return out


def _cellval(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# Coluna B..H do grid semanal = DOM..SÁB. `datetime.weekday()` é Seg=0..Dom=6;
# `(wd + 1) % 7` reindexa pra Dom=0 (mesma convenção de `_DOW`/`_build_weeks`).
# +1 porque a coluna A (índice 0) é o rótulo da linha, não um dia.
def _expected_day1_col(year: int, mm: int) -> int:
    return ((calendar.weekday(year, mm, 1) + 1) % 7) + 1


def _day1_col(rows: list) -> int | None:
    """Acha a coluna (1..7) onde o dia 1 aparece na 1ª linha de números do
    grid. Sinal determinístico de calendário — nome de aba e célula-título
    da planilha são comprovadamente mentirosos (ex.: aba 'CL JANEIRO' com
    título 'NOVEMBRO'); o layout dos dias não mente."""
    for row in rows:
        if not row:
            continue
        for c in range(1, 8):
            v = row[c] if c < len(row) else None
            if isinstance(v, (int, float)) and float(v).is_integer() \
                    and int(v) == 1:
                return c
    return None


def _max_day(rows: list) -> int:
    """Maior número de dia presente no grid (1..31). 2ª âncora: comprovar o
    MÊS, não só o ano — pega aba de mês trocado cujo dia-1 cai na mesma
    coluna por coincidência (ex.: novembro/30d disfarçado de janeiro/31d)."""
    mx = 0
    for row in rows:
        if not row:
            continue
        for c in range(1, 8):
            v = row[c] if c < len(row) else None
            if isinstance(v, (int, float)) and float(v).is_integer() \
                    and 1 <= int(v) <= 31:
                mx = max(mx, int(v))
    return mx


def _verify_2026(rows: list, sheet: str, yy: int, mm: int) -> None:
    """Recusa ingestão de aba que não **prove** ser {mm}/2026.

    A `ingest_xlsx` antiga confiava cegamente no `year` do chamador e só
    bloqueava year!=2026 — carimbando 2026 em abas 2025/sem-ano (bug que
    fabricou 8 meses falsos no teste-io). Agora a planilha tem que provar:
      1. ano no nome da aba (Briefing Design tem) ≠ 2026 → recusa imediata;
      2. âncora de calendário: a coluna do dia-1 no grid tem que casar com
         `weekday(2026, mm, 1)`. Não casou → recusa (e diz quais anos casam).
    """
    name_year = re.search(r"20\d\d", sheet)
    if name_year and int(name_year.group(0)) != INGEST_YEAR:
        raise GridError(
            f"Aba '{sheet}' tem ano {name_year.group(0)} no nome — ingestão "
            f"limitada a {INGEST_YEAR} (Fase 1). Recusada.")
    got = _day1_col(rows)
    if got is None:
        raise GridError(
            f"Aba '{sheet}': não achei a linha de números do grid pra provar "
            f"o ano. Ingestão {INGEST_YEAR}-only recusa o que não prova.")
    want = _expected_day1_col(yy, mm)
    if got != want:
        matches = [y for y in range(yy - 3, yy + 3)
                   if _expected_day1_col(y, mm) == got]
        raise GridError(
            f"Aba '{sheet}': o layout de dias do mês {mm:02d} põe o dia 1 na "
            f"coluna {got}, mas {yy}-{mm:02d} exige coluna {want}. Esta aba "
            f"corresponde a {matches or 'nenhum ano próximo'}, não a {yy}. "
            f"Ingestão {INGEST_YEAR}-only recusou (nome/título de aba são "
            f"não-confiáveis; calendário é a prova).")
    ndays = calendar.monthrange(yy, mm)[1]
    got_max = _max_day(rows)
    if got_max != ndays:
        raise GridError(
            f"Aba '{sheet}': o grid vai até o dia {got_max}, mas "
            f"{yy}-{mm:02d} tem {ndays} dias. Mês trocado (provável aba de "
            f"outro mês com dia-1 coincidente) — ingestão {INGEST_YEAR}-only "
            f"recusou. Falha-alto é o lado seguro: não regrava com mês errado.")


def _assert_no_mojibake(grid: dict, sheet: str) -> None:
    """Falha-alto se algum valor ingerido tiver U+FFFD (caractere de
    substituição). O parser in-process preserva UTF-8; FFFD só aparece se a
    ingestão foi conduzida por console/round-trip cp1252 (origem do `�` no
    teste-io). Nunca persistir mojibake silencioso."""
    blob = json.dumps(grid, ensure_ascii=False)
    if "�" in blob:
        raise GridError(
            f"Aba '{sheet}': valor ingerido contém U+FFFD (mojibake). A "
            f"ingestão tem que rodar in-process — nunca via console/`python "
            f"-c` capturado (cp1252 corrompe acento). Recusado sem gravar.")


def ingest_xlsx(path: str, *, sheet: str, brand: str, month,
                year: int = INGEST_YEAR, lib_dir: Path | None = None) -> dict:
    """
    Ingere UMA aba de uma planilha histórica → grid canônico (marca-mês).

    Escopo travado (Lucas 2026-05-18): só **2026**. `year` != 2026 levanta —
    o histórico 2025 fica fora; já valida o modelo com o ano corrente.

    `sheet`, `brand`, `month` e `year` são EXPLÍCITOS, MAS não confiados: a
    aba tem que **provar** ser {month}/2026 (`_verify_2026` — ano no nome
    quando há + âncora de calendário do dia-1). Aba que não prova é RECUSADA,
    nunca regravada com ano errado. Use `xlsx_sheets()` p/ escolher.
    Detecta o layout pelo rótulo da coluna A das linhas:
      - "ABORDAGEM" → Estratégia Mensal (PRODUTO + ABORDAGEM por dia);
      - "STORY"     → Briefing Design (PRODUTO + bloco STORY por dia).
    Mapeia colunas B..H = DOMINGO..SÁBADO; números de dia viram datas.
    """
    if int(year) != INGEST_YEAR:
        raise GridError(
            f"Ingestão limitada a {INGEST_YEAR} (Fase 1). year={year} fora "
            f"de escopo — o histórico 2025 não entra.")
    m = _norm_month(month, year)
    yy, mm = int(m[:4]), int(m[5:7])
    openpyxl = _load_openpyxl()
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise GridError(
                f"Aba '{sheet}' não existe em {path}. "
                f"Abas: {', '.join(wb.sheetnames)}")
        ws = wb[sheet]
        rows = [[c.value for c in row] for row in
                ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60),
                             max_col=26)]
    finally:
        wb.close()

    # Prova que a aba É {mm}/2026 antes de mapear qualquer dia (senão um
    # maio/2025 vira maio/2026 silenciosamente — o bug do teste-io).
    _verify_2026(rows, sheet, yy, mm)

    grid = new_grid(brand, m, focus_products=None, lib_dir=lib_dir,
                    save=False)
    by_date = {d["date"]: d for w in grid["weeks"] for d in w["days"]}

    def col_to_date(ci: int, day_num: int):
        try:
            return datetime(yy, mm, day_num).strftime("%Y-%m-%d")
        except ValueError:
            return None

    n = len(rows)
    for i in range(n):
        row = rows[i]
        label = _cellval(row[0]).upper() if row else ""
        if label not in ("PRODUTO",):
            continue
        # Linha PRODUTO: achar a linha de NÚMEROS logo acima (1-2 linhas).
        daynum_row = None
        for back in (1, 2):
            if i - back >= 0:
                cand = rows[i - back]
                nums = [(_cellval(cand[c])) for c in range(1, 8)]
                if any(x.isdigit() for x in nums):
                    daynum_row = cand
                    break
        if daynum_row is None:
            continue
        product_row = row
        # Linha de baixo: ABORDAGEM (estratégia) ou STORY (briefing design).
        kind_row = rows[i + 1] if i + 1 < n else None
        kind_label = _cellval(kind_row[0]).upper() if kind_row else ""
        for c in range(1, 8):  # B..H
            dn = _cellval(daynum_row[c])
            if not dn.isdigit():
                continue
            date_iso = col_to_date(c, int(dn))
            if not date_iso or date_iso not in by_date:
                continue
            day = by_date[date_iso]
            prod = _cellval(product_row[c])
            if prod:
                day["subject"] = prod
            if kind_label == "ABORDAGEM":
                ab = _cellval(kind_row[c])
                if ab:
                    day["approach"] = ab.upper()
            elif kind_label == "STORY":
                day["channel"] = "story"
                story = _parse_story(kind_row[c])
                if story:
                    day["lettering"] = story

    grid["ingestedFrom"] = {
        "file": Path(path).name, "sheet": sheet,
        "at": lc.now(),
    }
    _assert_no_mojibake(grid, sheet)
    return save_grid(grid, lib_dir=lib_dir)


# --------------------------------------------------------------------------- #
# Fase 2 — briefing → grid (mecânico/determinístico; o agente julga depois)
#
# Boundary com analyze-briefing é o dict `brief`:
#   {brand, month, launches[{date, product, label?, important?}],
#    focusProducts[slug], globalContent[{date?, note}], directionalNotes?}
# Validado por _validate_brief antes de consumir.
# --------------------------------------------------------------------------- #
_CAL_DATE_RE = re.compile(r"^\s*(\d{4})-(\d{2})(?:-(\d{2}))?\s*$")


def _read_seed_or_workspace_md(filename: str, ws_dir: Path,
                                seed_dir: Path) -> tuple[str | None, str | None]:
    """Lê um Markdown, preferindo workspace. Retorna (texto, caminho-usado).
    None/None se não existe em nenhum dos dois."""
    ws_path = ws_dir / filename
    try:
        return ws_path.read_text(encoding="utf-8"), str(ws_path)
    except OSError:
        pass
    seed_path = seed_dir / filename
    try:
        return seed_path.read_text(encoding="utf-8"), str(seed_path)
    except OSError:
        return None, None


def parse_calendar(year, *, lib_dir: Path | None = None) -> dict:
    """Lê `grids/calendar/<ano>.md` (fallback `calendar-seed/<ano>.md`).

    Tolerante a edição humana: linhas que não casam viram `warnings` com nº de
    linha — NUNCA silenciar (lição 9295e3b: verde sintético ≠ correto).

    Suporta:
      - `AAAA-MM-DD` (dia específico) → scope='day'
      - `AAAA-MM (mês)` ou `AAAA-MM (mes)` → scope='month' (datas tipo
        Outubro Rosa que valem o mês inteiro)

    Retorna {items: [{date, name, hook, scope}], warnings: [...], path}.
    Falha-alto se o arquivo `<ano>.md` não existe em workspace nem seed.
    """
    yy = int(year)
    if lib_dir is None:
        lib_dir = find_library_dir(create=False) or _ensure_ready(None)
    else:
        lib_dir = Path(lib_dir)
    _, cal_dir, *_ = _subdirs(lib_dir)
    text, used = _read_seed_or_workspace_md(
        f"{yy}.md", cal_dir, _CALENDAR_SEED_DIR)
    if text is None:
        raise GridError(
            f"Calendário '{yy}.md' não existe nem em {cal_dir} nem em "
            f"{_CALENDAR_SEED_DIR}. Crie em "
            f"grids/calendar/{yy}.md ou copie do seed (rebootstrap).")

    items: list[dict] = []
    warnings: list[str] = []
    in_table = False
    for ln, raw in enumerate(text.splitlines(), start=1):
        line = raw.strip()
        if not line or line.startswith("#") or line.startswith(">"):
            continue
        if not line.startswith("|"):
            continue
        # ignora a linha de cabeçalho e a separadora
        cells = [c.strip() for c in line.strip("|").split("|")]
        if len(cells) < 3:
            warnings.append(f"linha {ln}: <3 colunas — ignorada")
            continue
        # cabeçalho
        if cells[0].lower() in ("data", "date"):
            in_table = True
            continue
        # separador `---`
        if all(set(c.replace("-", "").replace(":", "").strip()) <= {""} for c in cells):
            continue
        if not in_table:
            # tabela sem cabeçalho reconhecível — tenta processar mesmo assim
            in_table = True

        date_cell, name, hook = cells[0], cells[1], cells[2]
        # `AAAA-MM (mês)` / `AAAA-MM (mes)` → month-wide
        mm_wide = re.match(r"^\s*(\d{4})-(\d{2})\s*\(\s*m[eê]s\s*\)\s*$",
                           date_cell, re.IGNORECASE)
        if mm_wide:
            y, m = int(mm_wide.group(1)), int(mm_wide.group(2))
            if y != yy:
                warnings.append(
                    f"linha {ln}: ano {y} no item '{name}' não bate com "
                    f"calendário {yy} — ignorada")
                continue
            items.append({"date": f"{y:04d}-{m:02d}",
                          "name": name, "hook": hook, "scope": "month"})
            continue
        m = _CAL_DATE_RE.match(date_cell)
        if not m or not m.group(3):
            warnings.append(
                f"linha {ln}: data '{date_cell}' fora de "
                f"'AAAA-MM-DD' ou 'AAAA-MM (mês)' — ignorada")
            continue
        y, mm, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y != yy:
            warnings.append(
                f"linha {ln}: ano {y} no item '{name}' não bate com "
                f"calendário {yy} — ignorada")
            continue
        try:
            iso = datetime(y, mm, dd).strftime("%Y-%m-%d")
        except ValueError as e:
            warnings.append(f"linha {ln}: data inválida ({e}) — ignorada")
            continue
        items.append({"date": iso, "name": name, "hook": hook, "scope": "day"})

    return {"items": items, "warnings": warnings, "path": used}


def read_rules(brand, *, lib_dir: Path | None = None) -> dict:
    """Texto integral de `grids/rules/<slug>.md` (fallback `rules-seed/`).

    NÃO PARSEIA — linguagem humana é de propósito (D4 do plano Fase 2). O
    agente lê o texto e aplica julgamento. Marca sem regras → missing=True.
    """
    if lib_dir is None:
        lib_dir = find_library_dir(create=False) or _ensure_ready(None)
    else:
        lib_dir = Path(lib_dir)
    rules_dir, *_ = _subdirs(lib_dir)
    slug = slugify(brand)
    text, used = _read_seed_or_workspace_md(
        f"{slug}.md", rules_dir, _RULES_SEED_DIR)
    return {"text": text, "missing": text is None, "path": used}


# --------------------------------------------------------------------------- #
# brief — validação (boundary com analyze-briefing)
# --------------------------------------------------------------------------- #
def _validate_brief(brief: dict, *, lib_dir: Path | None = None) -> dict:
    """Normaliza e valida o dict `brief`. Falha-alto em brand/month; reporta
    slug fantasma (produto inexistente) em `missing`, sem falhar (mesmo
    padrão de product_library.brand_from_briefing).

    Retorna {brief: <normalized>, missing: [...]}.
    """
    if not isinstance(brief, dict):
        raise InvalidGrid("brief não é um objeto.")
    brand = (brief.get("brand") or "").strip()
    if not brand:
        raise InvalidGrid("brief sem 'brand'.")
    month = brief.get("month")
    if not month:
        raise InvalidGrid("brief sem 'month'.")
    try:
        month_norm = _norm_month(month)
    except InvalidGrid:
        raise
    bslug = slugify(brand)

    # Catálogo de produtos da marca (pra reportar slugs fantasma sem falhar).
    # Import lazy: o core não quer dependência rígida de product_library.
    try:
        import product_library as pl
        known = {p.get("slug") for p in pl.list_products(bslug)}
    except Exception:
        known = set()

    missing: list[str] = []
    launches_norm: list[dict] = []
    for i, lau in enumerate(brief.get("launches", []) or []):
        if not isinstance(lau, dict):
            missing.append(f"launches[{i}]: não é objeto — ignorado")
            continue
        date = (lau.get("date") or "").strip()
        prod = (lau.get("product") or "").strip()
        if not date:
            missing.append(f"launches[{i}]: sem 'date'")
            continue
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date):
            missing.append(f"launches[{i}]: date '{date}' fora de "
                           f"AAAA-MM-DD")
            continue
        if prod and known and prod not in known:
            missing.append(
                f"launches[{i}].product '{prod}' não existe em "
                f"product-catalog/{bslug} — registre antes ou ajuste o slug")
        launches_norm.append({
            "date": date,
            "product": prod or None,
            "label": (lau.get("label") or "").strip() or None,
            "important": bool(lau.get("important", False)),
        })

    focus = []
    for i, p in enumerate(brief.get("focusProducts", []) or []):
        s = str(p).strip()
        if not s:
            continue
        if known and s not in known:
            missing.append(
                f"focusProducts[{i}] '{s}' não existe em "
                f"product-catalog/{bslug} — registre antes ou ajuste o slug")
        focus.append(s)

    global_content = []
    for i, g in enumerate(brief.get("globalContent", []) or []):
        if not isinstance(g, dict):
            continue
        date = (g.get("date") or "") or None
        if date and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(date)):
            missing.append(
                f"globalContent[{i}].date '{date}' fora de AAAA-MM-DD — "
                f"tratado como sem-data")
            date = None
        note = (g.get("note") or "").strip()
        if not note:
            continue
        global_content.append({"date": date, "note": note})

    return {
        "brief": {
            "brand": bslug,
            "month": month_norm,
            "launches": launches_norm,
            "focusProducts": focus,
            "globalContent": global_content,
            "directionalNotes": (brief.get("directionalNotes") or "").strip(),
        },
        "missing": missing,
    }


# --------------------------------------------------------------------------- #
# generate_from_briefing — andaime + âncoras + _slot tipado (mecânico)
# --------------------------------------------------------------------------- #
def _has_content(day: dict) -> bool:
    """Dia 'preenchido' = qualquer campo de _POST_FIELDS com valor não-vazio."""
    for k in _POST_FIELDS:
        v = day.get(k)
        if v is None or v == "" or v == {} or v == []:
            continue
        return True
    return False


def _existing_grid_has_content(lib_dir: Path, brand: str, month: str) -> bool:
    fp = _grid_path(lib_dir, brand, month)
    if not fp.is_file():
        return False
    try:
        existing = json.loads(fp.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return False
    # Seed marcado é exemplo descartável (não curadoria humana). Sem essa
    # checagem, o bootstrap colide com generate_from_briefing(brand=clinique,
    # month=2026-05) e a primeira UX da Fase 2 cai no overwrite-guard.
    if existing.get("_seed"):
        return False
    for w in existing.get("weeks", []):
        for d in w.get("days", []):
            if _has_content(d):
                return True
    return False


def _days_index(grid: dict) -> dict:
    return {d["date"]: d for w in grid.get("weeks", []) for d in w.get("days", [])}


def _date_add(date_iso: str, days: int) -> str:
    d = datetime.strptime(date_iso, "%Y-%m-%d")
    return datetime.fromordinal(d.toordinal() + days).strftime("%Y-%m-%d")


def generate_from_briefing(brief: dict, *, lib_dir: Path | None = None,
                           save: bool = True,
                           overwrite: bool = False,
                           min_posts: int = MIN_POSTS,
                           max_gap: int = MAX_GAP,
                           focus_intensity_days: int = FOCUS_INTENSITY_DAYS) -> dict:
    """Gera o ANDAIME do grid a partir do `brief` — mecânico, sem LLM, sem
    decidir conteúdo. O agente preenche product/subject/ref/etc. depois,
    via set_post, guiado por plan_card + rules/<marca>.md.

    Marca cada dia com `_slot = {kind, locked, hint, ...}` (taxonomia em
    `_SLOT_KINDS`); `_brief` persiste no top-level pra rastreabilidade.

    Idempotência segura: se já há grid <marca>/<mês> com qualquer dia não
    vazio e overwrite=False → InvalidGrid (não obliterar curadoria humana).
    """
    v = _validate_brief(brief, lib_dir=lib_dir)
    b = v["brief"]
    bslug, month = b["brand"], b["month"]
    yy, mm = int(month[:4]), int(month[5:7])

    lib_dir = _ensure_ready(lib_dir)
    if not overwrite and _existing_grid_has_content(lib_dir, bslug, month):
        raise InvalidGrid(
            f"Grid '{bslug}/{month}' já existe com conteúdo curado. "
            f"Passe overwrite=True para regenerar o andaime (sobrescreve "
            f"tudo) ou mova/edite à mão.")

    grid = new_grid(bslug, month, focus_products=b["focusProducts"],
                    lib_dir=lib_dir, save=False)
    by_date = _days_index(grid)
    ndays = calendar.monthrange(yy, mm)[1]

    # Calendário comemorativo do ano
    try:
        cal_data = parse_calendar(yy, lib_dir=lib_dir)
        cal_items = cal_data["items"]
    except GridError:
        cal_items = []

    # 1) Ancorar lançamentos + janela de intensidade
    for lau in b["launches"]:
        d0 = lau["date"]
        if d0 not in by_date:
            continue  # lançamento fora do mês — ignora silencioso (warning vai em audit)
        slot = {"kind": "launch_anchor", "locked": True,
                "hint": lau.get("label") or lau.get("product") or "lançamento"}
        if lau.get("product"):
            slot["product"] = lau["product"]
        by_date[d0]["_slot"] = slot
        # janela de intensidade nos dias seguintes (dentro do mês)
        for k in range(1, focus_intensity_days):
            d = _date_add(d0, k)
            if d not in by_date:
                break
            # não sobrepor uma âncora dura
            cur = by_date[d].get("_slot")
            if cur and cur.get("locked"):
                continue
            new_slot = {"kind": "launch_intensity", "locked": False,
                        "hint": lau.get("label") or lau.get("product")
                                or "intensidade pós-lançamento"}
            if lau.get("product"):
                new_slot["product"] = lau["product"]
            by_date[d]["_slot"] = new_slot

    # 2) Datas comemorativas do MÊS (scope='day' batendo o mês corrente).
    # Month-wide (scope='month') vai pro grid["_calendarMonth"] como dica
    # global pro plan-card; não amarra um dia específico.
    month_wide: list[dict] = []
    for item in cal_items:
        if item["scope"] == "day" and item["date"][:7] == month:
            d = item["date"]
            if d in by_date:
                cur = by_date[d].get("_slot")
                if cur and cur.get("locked"):
                    # âncora dura ganha; calendário entra como hint adicional
                    cur["calendarHook"] = f"{item['name']}: {item['hook']}"
                else:
                    by_date[d]["_slot"] = {
                        "kind": "calendar_hook", "locked": False,
                        "hint": f"{item['name']}: {item['hook']}",
                        "calendarHook": f"{item['name']}: {item['hook']}",
                    }
        elif item["scope"] == "month" and item["date"] == month:
            month_wide.append({"name": item["name"], "hook": item["hook"]})

    # 3) Intercalação no 4º/5º dia de qualquer foco contíguo
    # Detectado pelo produto que aparece em slots seguidos (launch_anchor +
    # launch_intensity ou focus_intercalation propagado). Lê o `product` do
    # _slot; quando ≥4 dias contíguos com mesmo produto, marca o seguinte
    # como `focus_intercalation` (proposta).
    dates_in_order = sorted(by_date.keys())
    run_prod = None
    run_len = 0
    for d in dates_in_order:
        slot = by_date[d].get("_slot")
        p = slot.get("product") if slot else None
        if p and p == run_prod:
            run_len += 1
            if run_len >= 4 and slot and not slot.get("locked") \
                    and slot["kind"] != "calendar_hook":
                by_date[d]["_slot"] = {
                    "kind": "focus_intercalation", "locked": False,
                    "hint": f"intercalar (foco '{run_prod}' há "
                            f"{run_len} dias) — outro foco/hero",
                    "previousFocus": run_prod,
                }
        else:
            run_prod = p
            run_len = 1 if p else 0

    # 4) Preencher slots vazios como hero_fill (proposta) e garantir cadência
    # (gap≤max_gap, posts≥min_posts). Cadência aqui é PROPOSTA — o conteúdo
    # do dia (product/subject) fica vazio: agente decide via set_post.
    for d in dates_in_order:
        if "_slot" not in by_date[d]:
            by_date[d]["_slot"] = {"kind": "hero_fill", "locked": False,
                                    "hint": "produto hero ou complementar"}

    # 5) Persistir _brief e dicas globais
    grid["_brief"] = {
        "launches": b["launches"],
        "focusProducts": b["focusProducts"],
        "globalContent": b["globalContent"],
        "directionalNotes": b["directionalNotes"],
        "calendarMonthWide": month_wide,
        "generatedAt": lc.now(),
        "validation": {"missing": v["missing"]},
    }
    grid["focusProducts"] = b["focusProducts"]

    if save:
        return save_grid(grid, lib_dir=lib_dir)
    return grid


# --------------------------------------------------------------------------- #
# audit_grid + plan_card (recomputáveis; não persistem)
# --------------------------------------------------------------------------- #
def _day_has_post(day: dict) -> bool:
    """Dia 'tem post' quando subject OU product está preenchido. Heurística
    deliberada — outras camadas (ref/lettering) podem existir sem post real."""
    return bool((day.get("subject") or "").strip()
                or (day.get("product") or "").strip())


def audit_grid(grid_or_brand, month=None, *,
               lib_dir: Path | None = None) -> dict:
    """Checa regras codificáveis sobre um grid concreto.

    Aceita o dict do grid OU (brand, month) — neste último caso chama
    get_grid. Mede só camada determinística (cobertura, gaps, datas
    comemorativas, focusProducts); julgamento (qual produto na data, hero)
    NÃO é auditado aqui — é follow-up pós-0.9.0.
    """
    if isinstance(grid_or_brand, dict):
        grid = grid_or_brand
    else:
        grid = get_grid(grid_or_brand, month, lib_dir=lib_dir)
    yy, mm = int(grid["month"][:4]), int(grid["month"][5:7])
    ndays_in_month = calendar.monthrange(yy, mm)[1]

    days_sorted = sorted(
        (d for w in grid.get("weeks", []) for d in w.get("days", [])),
        key=lambda d: d.get("date", ""))

    posts = 0
    products_present: set = set()
    dates_with_post: set = set()
    gap_runs: list[int] = []
    cur_gap = 0
    for d in days_sorted:
        if _day_has_post(d):
            posts += 1
            dates_with_post.add(d["date"])
            if d.get("product"):
                products_present.add(d["product"])
            if cur_gap > 0:
                gap_runs.append(cur_gap)
                cur_gap = 0
        else:
            cur_gap += 1
    if cur_gap > 0:
        gap_runs.append(cur_gap)
    max_gap = max(gap_runs) if gap_runs else 0

    # cobertura: dias com post / dias do mês presentes no grid
    days_in_grid = len(days_sorted)
    coverage = posts / days_in_grid if days_in_grid else 0.0

    # focusCoverage: para cada foco declarado em _brief.focusProducts (ou
    # grid.focusProducts como fallback), conta dias com aquele produto
    foci = ((grid.get("_brief") or {}).get("focusProducts")
            or grid.get("focusProducts") or [])
    focus_cov = {p: sum(1 for d in days_sorted if d.get("product") == p)
                 for p in foci}

    # datesPresent: datas comemorativas do MÊS que viraram post
    try:
        cal = parse_calendar(yy, lib_dir=lib_dir)["items"]
    except GridError:
        cal = []
    cal_days = [it for it in cal
                if it["scope"] == "day" and it["date"][:7] == grid["month"]]
    dates_present = {it["date"]: it["date"] in dates_with_post
                     for it in cal_days}
    dates_present_ratio = (
        sum(1 for v in dates_present.values() if v) / len(cal_days)
        if cal_days else 1.0)

    warnings: list[str] = []
    if posts < MIN_POSTS:
        warnings.append(f"posts={posts} < MIN_POSTS={MIN_POSTS}")
    if max_gap > MAX_GAP:
        warnings.append(f"maxGap={max_gap} > MAX_GAP={MAX_GAP}")
    if days_in_grid != ndays_in_month:
        warnings.append(
            f"dias no grid ({days_in_grid}) != dias do mês ({ndays_in_month})")
    for p, n in focus_cov.items():
        if n == 0:
            warnings.append(f"focusProduct '{p}' não aparece no grid")

    return {
        "brand": grid["brand"], "month": grid["month"],
        "posts": posts, "coverage": round(coverage, 3),
        "maxGap": max_gap, "gaps": gap_runs,
        "focusCoverage": focus_cov,
        "datesPresent": dates_present,
        "datesPresentRatio": round(dates_present_ratio, 3),
        "warnings": warnings,
    }


def plan_card(grid: dict, *, lib_dir: Path | None = None) -> dict:
    """Dossiê de julgamento pro agente: panorama + rules.md inline + slots
    a decidir. Recomputável (não persiste); chame de novo após set_post."""
    if not isinstance(grid, dict):
        raise InvalidGrid("plan_card precisa de um dict de grid.")
    brand = grid.get("brand", "")
    yy = int(grid["month"][:4])

    rules = read_rules(brand, lib_dir=lib_dir)
    try:
        cal = parse_calendar(yy, lib_dir=lib_dir)
    except GridError:
        cal = {"items": [], "warnings": [], "path": None}

    brief = grid.get("_brief") or {}
    slots_todo: list[dict] = []
    for w in grid.get("weeks", []):
        for d in w.get("days", []):
            slot = d.get("_slot")
            if not slot:
                continue
            if _day_has_post(d):
                continue  # já decidido pelo agente
            slots_todo.append({
                "date": d["date"], "dow": d["dow"], "_slot": slot,
                "currentContent": {k: d.get(k) for k in _POST_FIELDS
                                   if d.get(k) not in (None, "", {}, [])},
            })

    return {
        "brand": brand, "month": grid.get("month"),
        "rules": rules,
        "calendarWarnings": cal["warnings"],
        "launches": brief.get("launches", []),
        "focusProducts": brief.get("focusProducts", []),
        "globalContent": brief.get("globalContent", []),
        "directionalNotes": brief.get("directionalNotes", ""),
        "calendarMonthWide": brief.get("calendarMonthWide", []),
        "slotsTodo": slots_todo,
        "audit": audit_grid(grid, lib_dir=lib_dir),
    }


# --------------------------------------------------------------------------- #
# render
# --------------------------------------------------------------------------- #
def render_grids(lib_dir: Path | None = None,
                 template_path: Path | None = None) -> Path:
    """Lê o template, injeta {grids:[...]} (workspace ou seed) e escreve
    `<lib_dir>/grids.html`. Sempre derivado — nunca lido como dado."""
    lib_dir = _ensure_ready(lib_dir)
    grids = sorted(_resolve(lib_dir)[0], key=_sort_key)
    payload = json.dumps({"grids": grids}, ensure_ascii=False)
    tpl = Path(template_path or _TEMPLATE_FILE).read_text(encoding="utf-8")
    injected = lc.inject_placeholder(tpl, _PLACEHOLDER, payload, GridError)
    _, _, _, _, html_path = _subdirs(lib_dir)
    lc.atomic_write(html_path, injected)
    return html_path


def open_grids(lib_dir: Path | None = None) -> Path:
    """Regenera e devolve o caminho do grids.html (a skill manda abrir)."""
    return render_grids(lib_dir)
